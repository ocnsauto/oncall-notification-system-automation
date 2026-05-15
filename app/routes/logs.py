import io
from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, send_file, current_app
from flask_login import login_required
from app import db
from app.models import Incident, NotificationLog, Engineer
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logs_bp = Blueprint("logs", __name__, url_prefix="/logs")


@logs_bp.route("/")
@login_required
def index():
    page = request.args.get("page", 1, type=int)
    status_filter = request.args.get("status", "")
    engineer_filter = request.args.get("engineer_id", "", type=str)
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")

    query = Incident.query

    if status_filter:
        query = query.filter(Incident.status == status_filter)
    if date_from:
        try:
            query = query.filter(Incident.triggered_at >= datetime.strptime(date_from, "%Y-%m-%d"))
        except ValueError:
            pass
    if date_to:
        try:
            dt_to = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(Incident.triggered_at < dt_to)
        except ValueError:
            pass

    incidents = query.order_by(Incident.triggered_at.desc()).paginate(page=page, per_page=20)
    engineers = Engineer.query.order_by(Engineer.name).all()

    return render_template(
        "logs/index.html",
        incidents=incidents,
        engineers=engineers,
        status_filter=status_filter,
        engineer_filter=engineer_filter,
        date_from=date_from,
        date_to=date_to,
    )


@logs_bp.route("/export")
@login_required
def export():
    range_param = request.args.get("range", "weekly")
    now = datetime.utcnow()

    if range_param == "monthly":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        label = now.strftime("%Y-%m")
    else:
        start = now - timedelta(days=now.weekday(), hours=now.hour, minutes=now.minute)
        start = start.replace(second=0, microsecond=0)
        label = f"{now.strftime('%Y')}-W{now.strftime('%W')}"

    incidents = Incident.query.filter(Incident.triggered_at >= start).order_by(Incident.triggered_at).all()

    wb = openpyxl.Workbook()
    _build_summary_sheet(wb.active, incidents, label)
    detail_sheet = wb.create_sheet("Detail Log")
    _build_detail_sheet(detail_sheet, incidents)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"oncall-log-{label}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1a1a2e")
    cell.alignment = Alignment(horizontal="left")


def _build_summary_sheet(ws, incidents, label):
    ws.title = "Summary"
    headers = ["#", "Triggered At (GMT+8)", "Subject", "Body Snippet", "Call Attempts", "Status", "Engineer"]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell)
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 20

    from zoneinfo import ZoneInfo
    import datetime
    timezone_str = current_app.config.get("APP_TIMEZONE", "Asia/Manila")

    for i, inc in enumerate(incidents, 1):
        calls = [l for l in inc.notification_logs if l.type == "call"]
        status_list = []
        engineer_list = []
        for call in calls:
            if call.status == "completed":
                status_val = "Answered"
            elif call.status == "no-answer":
                status_val = "Timed Out"
            elif call.status == "busy":
                status_val = "Declined"
            elif call.status == "failed":
                status_val = "Failed"
            else:
                status_val = call.status or "—"
            status_list.append(status_val)
            eng_name = call.engineer.name if call.engineer else "Unknown"
            engineer_list.append(eng_name)

        status_str = "\n".join(status_list) if status_list else "—"
        engineer_str = "\n".join(engineer_list) if engineer_list else "—"

        if inc.triggered_at.tzinfo is None:
            dt_utc = inc.triggered_at.replace(tzinfo=datetime.timezone.utc)
        else:
            dt_utc = inc.triggered_at
        local_time = dt_utc.astimezone(ZoneInfo(timezone_str)).strftime("%Y-%m-%d %H:%M:%S")

        subject = (inc.email_subject or "")[:80]
        snippet = (inc.body_snippet or "")[:120]

        ws.append([
            i,
            local_time,
            subject,
            snippet,
            len(calls),
            status_str,
            engineer_str,
        ])

        row_idx = ws.max_row
        ws.cell(row=row_idx, column=6).alignment = Alignment(wrapText=True)
        ws.cell(row=row_idx, column=7).alignment = Alignment(wrapText=True)



def _build_detail_sheet(ws, incidents):
    headers = ["Log ID", "Incident ID", "Engineer", "Type", "Status", "Twilio SID", "Timestamp (GMT+8)"]
    ws.append(headers)
    for cell in ws[1]:
        _header_style(cell)
    ws.column_dimensions["G"].width = 22

    incident_ids = [inc.id for inc in incidents]
    if not incident_ids:
        return
    logs = (
        NotificationLog.query.filter(NotificationLog.incident_id.in_(incident_ids))
        .order_by(NotificationLog.timestamp)
        .all()
    )
    from zoneinfo import ZoneInfo
    import datetime
    timezone_str = current_app.config.get("APP_TIMEZONE", "Asia/Manila")

    for log in logs:
        if log.timestamp.tzinfo is None:
            dt_utc = log.timestamp.replace(tzinfo=datetime.timezone.utc)
        else:
            dt_utc = log.timestamp
        local_time = dt_utc.astimezone(ZoneInfo(timezone_str)).strftime("%Y-%m-%d %H:%M:%S")

        status_val = log.status or "—"
        if log.type == "call":
            if log.status == "completed":
                status_val = "Answered"
            elif log.status == "no-answer":
                status_val = "Timed Out"
            elif log.status == "busy":
                status_val = "Declined"
            elif log.status == "failed":
                status_val = "Failed"

        ws.append([
            log.id,
            log.incident_id,
            log.engineer.name if log.engineer else "—",
            log.type,
            status_val,
            log.twilio_sid or "—",
            local_time,
        ])
